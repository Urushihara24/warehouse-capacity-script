import requests
import pandas as pd
import os
import shutil
import time
import re
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Ошибка: Требуется библиотека openpyxl. Установите ее: py -m pip install openpyxl")
    exit()

try:
    import gspread
    from google.oauth2.service_account import Credentials
except ImportError:
    print("Ошибка: Требуются библиотеки gspread и google-auth. Установите их: pip install gspread google-auth-oauthlib")
    exit()

pd.set_option('future.no_silent_downcasting', True)

# --- ОСНОВНЫЕ НАСТРОЙКИ ---
LOGIN_API_URL = 'https://wms.wbwh.tech/srv/auth_phone_password/api/login'
DATA_URL = 'https://wms.wbwh.tech/srv/wms_front_reports_common_proxy_adapter/sql/wh_filling_by_stage_street_mx'
USERNAME = '+79966240570'
PASSWORD = 'UrushiharaHanzou23@@'

# --- НАСТРОЙКИ GOOGLE SHEETS ---
try:
    # Этот путь работает при запуске как .py файл
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    # Этот путь работает в интерактивных средах (Jupyter, IDLE и т.д.)
    SCRIPT_DIR = os.getcwd()

GSHEET_CRED_FILE = os.path.join(SCRIPT_DIR, 'credentials.json')
GSHEET_ID = '18eSznp_9Vim-uf6dlBbCxr8bYFbC_me7dgEn_NlWqn8' # ID вашей рабочей таблицы

WAREHOUSES_TO_PROCESS = [
    {'name': 'Основной_Склад_112', 'id': 112, 'floors_to_process': [1, 3, 4, 5, 6], 'rows_to_process': range(0, 27)},
    {'name': 'Казань_4_Питание', 'id': 291, 'floors_to_process': [2], 'rows_to_process': range(0, 27)}
]
BASE_DOWNLOAD_DIR = '.'

# --- ДАННЫЕ ПО ПОРОГАМ ЗАПОЛНЕННОСТИ ---
# ИЗМЕНЕНИЕ: Добавлены точные данные для категорий 2-го этажа
FILL_THRESHOLDS_PER_FLOOR = {
    'Общая_Площадка': {'Общее': {'100%': 7566893, '95%': 7158013}},
    1: {'Общее': {'100%': 927303, '95%': 880938},
        'Химия': {'100%': 444338, '95%': 422121},
        'Красота': {'100%': 333141, '95%': 316484},
        'КГТ': {'100%': 149824, '95%': 142333}},

    2: {'Общее': {'100%': 2412008, '95%': 2291408}, # Сумма вместимостей СГТ и МГТ
        'Полочный товар':   {'100%': 359248, '95%': 341286},  # Данные из вашей таблицы
        'Коробочный товар': {'100%': 2052760, '95%': 1950122}}, # Данные из вашей таблицы

    3: {'Общее': {'100%': 1362890, '95%': 1335578},
        'Полочный товар': {'100%': 256800, '95%': 133666},
        'Коробочный товар': {'100%': 1102100, '95%': 1201912}},
    4: {'Общее': {'100%': 1758900, '95%': 1600504},
        'Полочный товар': {'100%': 385200, '95%': 151813},
        'Коробочный товар': {'100%': 1373700, '95%': 1448691}},
    5: {'Общее': {'100%': 1758900, '95%': 1643834},
        'Полочный товар': {'100%': 385200, '95%': 1526003},
        'Коробочный товар': {'100%': 1373700, '95%': 117531}},
    6: {'Общее': {'100%': 1758900, '95%': 1697159},
        'Полочный товар': {'100%': 385200, '95%': 1586913},
        'Коробочный товар': {'100%': 1373700, '95%': 110246}}
}
FILL_THRESHOLDS_PER_ROW = {
    1: {0: {'100%': 21074, '95%': 20020, '85%': 17913}, 1: {'100%': 37996, '95%': 36096, '85%': 32297}, 2: {'100%': 37996, '95%': 36096, '85%': 32297}, 3: {'100%': 37996, '95%': 36096, '85%': 32297}, 4: {'100%': 37996, '95%': 36096, '85%': 32297}, 5: {'100%': 37996, '95%': 36096, '85%': 32297}, 6: {'100%': 37996, '95%': 36096, '85%': 32297}, 7: {'100%': 37996, '95%': 36096, '85%': 32297}, 8: {'100%': 37452, '95%': 35579, '85%': 31834}, 9: {'100%': 37996, '95%': 36096, '85%': 32297}, 10: {'100%': 37996, '95%': 36096, '85%': 32297}, 11: {'100%': 36798, '95%': 34958, '85%': 31278}, 12: {'100%': 36798, '95%': 34958, '85%': 31278}, 13: {'100%': 37996, '95%': 36096, '85%': 32297}, 14: {'100%': 37996, '95%': 36096, '85%': 32297}, 15: {'100%': 37996, '95%': 36096, '85%': 32297}, 16: {'100%': 37996, '95%': 36096, '85%': 32297}, 17: {'100%': 37996, '95%': 36096, '85%': 32297}, 18: {'100%': 37996, '95%': 36096, '85%': 32297}, 19: {'100%': 37996, '95%': 36096, '85%': 32297}, 20: {'100%': 37996, '95%': 36096, '85%': 32297}, 21: {'100%': 37996, '95%': 36096, '85%': 32297}, 22: {'100%': 37996, '95%': 36096, '85%': 32297}, 23: {'100%': 36107, '95%': 34302, '85%': 30691}, 24: {'100%': 34334, '95%': 32617, '85%': 29184}, 25: {'100%': 59252, '95%': 56289, '85%': 50364}, 26: {'100%': 1394, '95%': 1324, '85%': 1185}},
    2: {0: {'100%': 18400, '95%': 17480, '85%': 15640}, 1: {'100%': 105340, '95%': 100073, '85%': 89539}, 2: {'100%': 96220, '95%': 91409, '85%': 81787}, 3: {'100%': 96670, '95%': 91837, '85%': 82170}, 4: {'100%': 96670, '95%': 91837, '85%': 82170}, 5: {'100%': 96670, '95%': 91837, '85%': 82170}, 6: {'100%': 96670, '95%': 91837, '85%': 82170}, 7: {'100%': 96670, '95%': 91837, '85%': 82170}, 8: {'100%': 89080, '95%': 84626, '85%': 75718}, 9: {'100%': 89080, '95%': 84626, '85%': 75718}, 10: {'100%': 96670, '95%': 91837, '85%': 82170}, 11: {'100%': 93900, '95%': 89205, '85%': 79815}, 12: {'100%': 93900, '95%': 89205, '85%': 79815}, 13: {'100%': 96670, '95%': 91837, '85%': 82170}, 14: {'100%': 96670, '95%': 91837, '85%': 82170}, 15: {'100%': 96670, '95%': 91837, '85%': 82170}, 16: {'100%': 96670, '95%': 91837, '85%': 82170}, 17: {'100%': 96670, '95%': 91837, '85%': 82170}, 18: {'100%': 96670, '95%': 91837, '85%': 82170}, 19: {'100%': 96670, '95%': 91837, '85%': 82170}, 20: {'100%': 96670, '95%': 91837, '85%': 82170}, 21: {'100%': 96670, '95%': 91837, '85%': 82170}, 22: {'100%': 96670, '95%': 91837, '85%': 82170}, 23: {'100%': 93510, '95%': 88835, '85%': 79484}, 24: {'100%': 65868, '95%': 62575, '85%': 55988}, 25: {'100%': 98940, '95%': 93993, '85%': 84099}, 26: {'100%': 17220, '95%': 16359, '85%': 14637}},
    3: {0: {'100%': 11058, '95%': 10505, '85%': 9399}, 1: {'100%': 55558, '95%': 52780, '85%': 47224}, 2: {'100%': 54308, '95%': 51593, '85%': 46162}, 3: {'100%': 54308, '95%': 51593, '85%': 46162}, 4: {'100%': 54308, '95%': 51593, '85%': 46162}, 5: {'100%': 54308, '95%': 51593, '85%': 46162}, 6: {'100%': 54308, '95%': 51593, '85%': 46162}, 7: {'100%': 54308, '95%': 51593, '85%': 46162}, 8: {'100%': 50308, '95%': 47793, '85%': 42762}, 9: {'100%': 54308, '95%': 51593, '85%': 46162}, 10: {'100%': 54308, '95%': 51593, '85%': 46162}, 11: {'100%': 52558, '95%': 49930, '85%': 44674}, 12: {'100%': 52558, '95%': 49930, '85%': 44674}, 13: {'100%': 54308, '95%': 51593, '85%': 46162}, 14: {'100%': 54308, '95%': 51593, '85%': 46162}, 15: {'100%': 54308, '95%': 51593, '85%': 46162}, 16: {'100%': 54308, '95%': 51593, '85%': 46162}, 17: {'100%': 54308, '95%': 51593, '85%': 46162}, 18: {'100%': 54308, '95%': 51593, '85%': 46162}, 19: {'100%': 54308, '95%': 51593, '85%': 46162}, 20: {'100%': 54308, '95%': 51593, '85%': 46162}, 21: {'100%': 54308, '95%': 51593, '85%': 46162}, 22: {'100%': 54308, '95%': 51593, '85%': 46162}, 23: {'100%': 52558, '95%': 49930, '85%': 44674}, 24: {'100%': 47058, '95%': 44705, '85%': 39999}, 25: {'100%': 52632, '95%': 50000, '85%': 44737}, 26: {'100%': 11058, '95%': 10505, '85%': 9399}},
    4: {0: {'100%': 18366, '95%': 17448, '85%': 15611}, 1: {'100%': 71394, '95%': 67824, '85%': 60685}, 2: {'100%': 69764, '95%': 66276, '85%': 59299}, 3: {'100%': 69764, '95%': 66276, '85%': 59299}, 4: {'100%': 69764, '95%': 66276, '85%': 59299}, 5: {'100%': 69764, '95%': 66276, '85%': 59299}, 6: {'100%': 69764, '95%': 66276, '85%': 59299}, 7: {'100%': 69764, '95%': 66276, '85%': 59299}, 8: {'100%': 64548, '95%': 61321, '85%': 54866}, 9: {'100%': 69764, '95%': 66276, '85%': 59299}, 10: {'100%': 69764, '95%': 66276, '85%': 59299}, 11: {'100%': 67482, '95%': 64108, '85%': 57360}, 12: {'100%': 67482, '95%': 64108, '85%': 57360}, 13: {'100%': 69764, '95%': 66276, '85%': 59299}, 14: {'100%': 69764, '95%': 66276, '85%': 59299}, 15: {'100%': 69764, '95%': 66276, '85%': 59299}, 16: {'100%': 69764, '95%': 66276, '85%': 59299}, 17: {'100%': 69764, '95%': 66276, '85%': 59299}, 18: {'100%': 69764, '95%': 66276, '85%': 59299}, 19: {'100%': 69764, '95%': 66276, '85%': 59299}, 20: {'100%': 69764, '95%': 66276, '85%': 59299}, 21: {'100%': 69764, '95%': 66276, '85%': 59299}, 22: {'100%': 69764, '95%': 66276, '85%': 59299}, 23: {'100%': 66178, '95%': 62869, '85%': 56251}, 24: {'100%': 60310, '95%': 57295, '85%': 51264}, 25: {'100%': 69022, '95%': 65571, '85%': 58669}, 26: {'100%': 18366, '95%': 17448, '85%': 15611}},
    5: {0: {'100%': 18366, '95%': 17448, '85%': 15611}, 1: {'100%': 71394, '95%': 67824, '85%': 60685}, 2: {'100%': 69764, '95%': 66276, '85%': 59299}, 3: {'100%': 69764, '95%': 66276, '85%': 59299}, 4: {'100%': 69764, '95%': 66276, '85%': 59299}, 5: {'100%': 69764, '95%': 66276, '85%': 59299}, 6: {'100%': 69764, '95%': 66276, '85%': 59299}, 7: {'100%': 69764, '95%': 66276, '85%': 59299}, 8: {'100%': 64548, '95%': 61321, '85%': 54866}, 9: {'100%': 69764, '95%': 66276, '85%': 59299}, 10: {'100%': 69764, '95%': 66276, '85%': 59299}, 11: {'100%': 67482, '95%': 64108, '85%': 57360}, 12: {'100%': 67482, '95%': 64108, '85%': 57360}, 13: {'100%': 69764, '95%': 66276, '85%': 59299}, 14: {'100%': 69764, '95%': 66276, '85%': 59299}, 15: {'100%': 69764, '95%': 66276, '85%': 59299}, 16: {'100%': 69764, '95%': 66276, '85%': 59299}, 17: {'100%': 69764, '95%': 66276, '85%': 59299}, 18: {'100%': 69764, '95%': 66276, '85%': 59299}, 19: {'100%': 69764, '95%': 66276, '85%': 59299}, 20: {'100%': 69764, '95%': 66276, '85%': 59299}, 21: {'100%': 69764, '95%': 66276, '85%': 59299}, 22: {'100%': 69764, '95%': 66276, '85%': 59299}, 23: {'100%': 66178, '95%': 62869, '85%': 56251}, 24: {'100%': 60310, '95%': 57295, '85%': 51264}, 25: {'100%': 69022, '95%': 65571, '85%': 58669}, 26: {'100%': 18366, '95%': 17448, '85%': 15611}},
    6: {0: {'100%': 18366, '95%': 17448, '85%': 15611}, 1: {'100%': 71394, '95%': 67824, '85%': 60685}, 2: {'100%': 69764, '95%': 66276, '85%': 59299}, 3: {'100%': 69764, '95%': 66276, '85%': 59299}, 4: {'100%': 69764, '95%': 66276, '85%': 59299}, 5: {'100%': 69764, '95%': 66276, '85%': 59299}, 6: {'100%': 69764, '95%': 66276, '85%': 59299}, 7: {'100%': 69764, '95%': 66276, '85%': 59299}, 8: {'100%': 64548, '95%': 61321, '85%': 54866}, 9: {'100%': 69764, '95%': 66276, '85%': 59299}, 10: {'100%': 69764, '95%': 66276, '85%': 59299}, 11: {'100%': 67482, '95%': 64108, '85%': 57360}, 12: {'100%': 67482, '95%': 64108, '85%': 57360}, 13: {'100%': 69764, '95%': 66276, '85%': 59299}, 14: {'100%': 69764, '95%': 66276, '85%': 59299}, 15: {'100%': 69764, '95%': 66276, '85%': 59299}, 16: {'100%': 69764, '95%': 66276, '85%': 59299}, 17: {'100%': 69764, '95%': 66276, '85%': 59299}, 18: {'100%': 69764, '95%': 66276, '85%': 59299}, 19: {'100%': 69764, '95%': 66276, '85%': 59299}, 20: {'100%': 69764, '95%': 66276, '85%': 59299}, 21: {'100%': 69764, '95%': 66276, '85%': 59299}, 22: {'100%': 69764, '95%': 66276, '85%': 59299}, 23: {'100%': 66178, '95%': 62869, '85%': 56251}, 24: {'100%': 60310, '95%': 57295, '85%': 51264}, 25: {'100%': 69022, '95%': 65571, '85%': 58669}, 26: {'100%': 18366, '95%': 17448, '85%': 15611}}
}

# --- ЦВЕТОВЫЕ СХЕМЫ И ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ---
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

def get_status_by_thresholds(current_value, thresholds):
    if not thresholds: return "НЕТ ДАННЫХ", 0, None
    val_100, val_95, val_85 = thresholds.get('100%'), thresholds.get('95%'), thresholds.get('85%')
    if val_100 is None or val_95 is None: return "НЕТ ДАННЫХ", 0, None
    if val_85 is None: val_85 = val_100 * 0.85
    if current_value >= val_95: status, color_fill = "🔴 КРИТИЧЕСКИЙ УРОВЕНЬ", RED_FILL
    elif current_value >= val_85: status, color_fill = "🟡 ВЫСОКАЯ ЗАГРУЗКА", YELLOW_FILL
    else: status, color_fill = "🟢 СВОБОДНО", GREEN_FILL
    percentage = (current_value / val_100 * 100) if val_100 > 0 else 0
    return status, percentage, color_fill

def login_and_get_token(session):
    payload = {'phone_number': USERNAME, 'password': PASSWORD}
    print("Отправка запроса на авторизацию...")
    try:
        response = session.post(LOGIN_API_URL, json=payload, timeout=10)
        response.raise_for_status()
        access_token = response.json().get('data', {}).get('access_token')
        if not access_token: raise ValueError("Токен не получен.")
        print("Авторизация прошла успешно!")
        return access_token
    except requests.exceptions.RequestException as e: print(f"Ошибка при авторизации: {e}"); raise

def download_and_save_data(session, token, warehouse_config, output_dir):
    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    os.makedirs(output_dir, exist_ok=True)
    wh_id = warehouse_config['id']
    print(f"\n--- Скачивание данных для склада ID: {wh_id} ({warehouse_config['name']}) ---")
    for floor in warehouse_config['floors_to_process']:
        for row in warehouse_config['rows_to_process']:
            try:
                payload = {'wh_id': wh_id, 'stage': floor, 'street': row}
                print(f"  Запрос: Этаж {floor}, Ряд {row}...", end='', flush=True)
                response = session.post(DATA_URL, json=payload, headers=headers, timeout=20)
                response.raise_for_status()
                data = response.json()
                if data and isinstance(data, list):
                    df = pd.DataFrame(data)
                    df.to_excel(os.path.join(output_dir, f'floor-{floor}_row-{row}.xlsx'), index=False)
                    print(f" ✓ Cохранено {len(data)} строк.")
                else: print(" Пусто.")
            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 401: print(" ОШИБКА 401: Токен недействителен."); raise
                print(f" ✗ ОШИБКА HTTP: {e.response.status_code}")
            except Exception as e: print(f" ✗ Критическая ошибка: {e}")
            time.sleep(0.5)
    print(f"--- Скачивание для склада ID: {wh_id} завершено ---")

def analyze_warehouse_data(directory):
    print(f"Анализ данных в '{directory}'...")
    try:
        all_files = [os.path.join(path, name) for path, _, files in os.walk(directory) for name in files if name.endswith('.xlsx')]
        if not all_files:
            print(f"В папке '{directory}' не найдены файлы для анализа.")
            return {}, []
    except FileNotFoundError:
        print(f"Папка '{directory}' не найдена.")
        return {}, []

    category_data, row_data_for_excel = {}, []
    
    for file in sorted(all_files):
        basename = os.path.basename(file)
        match = re.search(r'floor-(\d+)_row-(\d+)\.xlsx', basename)
        if not match: continue
        floor, row_num = int(match.group(1)), int(match.group(2))
        try:
            df = pd.read_excel(file, engine='openpyxl')
            
            original_columns = df.columns
            df.columns = [str(c).strip().lower() for c in original_columns]
            
            qty_col = next((c for c in df.columns if c in ['qty', 'кол-во']), None)
            box_id_col = next((c for c in df.columns if c in ['box_id', 'коробка']), None)
            place_col = next((c for c in df.columns if c in ['place_name', 'place_id', 'мх']), None)

            if not qty_col: continue
            df[qty_col] = pd.to_numeric(df[qty_col], errors='coerce').fillna(0).astype(int)
            if box_id_col: df[box_id_col] = pd.to_numeric(df[box_id_col], errors='coerce').fillna(0).astype(int)
            if place_col: df[place_col] = df[place_col].astype(str)

            category_data.setdefault(floor, {})
            
            if floor == 1:
                if not place_col: continue
                category_data[floor].setdefault('Химия', {'СГТ': 0, 'МГТ': 0}); category_data[floor].setdefault('Красота', {'СГТ': 0, 'МГТ': 0}); category_data[floor].setdefault('КГТ', {'Общее': 0}); category_data[floor].setdefault('Неопределенная категория', {'Общее': 0})
                for _, row in df.iterrows():
                    qty, box_id, place_str = row[qty_col], row.get(box_id_col, 0), row[place_col]
                    
                    try:
                        parts = place_str.split('.')
                        stellazh = -1 
                        if len(parts) == 6 and parts[3].isdigit():
                            stellazh = int(parts[3])
                        elif len(parts) == 5 and parts[2].isdigit():
                            stellazh = int(parts[1])
                    except (ValueError, IndexError):
                        stellazh = -1
                        
                    categorized = False
                    if 0 <= row_num <= 15 and 1 <= stellazh <= 64: category_data[floor]['Химия']['СГТ' if box_id == 0 else 'МГТ'] += qty; categorized = True
                    elif 16 <= row_num <= 25 and 1 <= stellazh <= 64: category_data[floor]['Красота']['СГТ' if box_id == 0 else 'МГТ'] += qty; categorized = True
                    elif (1 <= row_num <= 25 and 69 <= stellazh <= 222) or (row_num == 26 and 1 <= stellazh <= 42): category_data[floor]['КГТ']['Общее'] += qty; categorized = True
                    if not categorized: category_data[floor]['Неопределенная категория']['Общее'] += qty
            else:
                category_data[floor].setdefault('Полочный товар', {'Общее': 0}); category_data[floor].setdefault('Коробочный товар', {'Общее': 0})
                if not box_id_col: category_data[floor]['Полочный товар']['Общее'] += df[qty_col].sum()
                else: category_data[floor]['Полочный товар']['Общее'] += df[df[box_id_col] == 0][qty_col].sum(); category_data[floor]['Коробочный товар']['Общее'] += df[df[box_id_col] != 0][qty_col].sum()
            
            current_total_row = df[qty_col].sum()
            if current_total_row > 0:
                row_thresholds = FILL_THRESHOLDS_PER_ROW.get(floor, {}).get(row_num)
                if row_thresholds:
                    row_status, row_percent, row_color = get_status_by_thresholds(current_total_row, row_thresholds)
                    items_to_95 = max(0, int(row_thresholds['95%'] - current_total_row)) if row_thresholds and '95%' in row_thresholds else "N/A"
                    row_data_for_excel.append({'Этаж': floor, 'Ряд': row_num, 'Фактическое кол-во ШК': current_total_row, 'Макс. вместимость (100%)': (row_thresholds or {}).get('100%'), '% заполнения': row_percent, 'Осталось ШК до 95%': items_to_95, 'Статус': row_status, 'Цвет': row_color})
        except Exception as e: print(f"  Критическая ошибка при обработке файла {file}: {e}")
    
    return category_data, row_data_for_excel

def generate_summary_txt_report(report_name, category_data):
    txt_filename = f'summary_report_{report_name}_{datetime.now().strftime("%Y-%m-%d")}.txt'
    report_lines = ["*"*60, f"Сводный отчет по категориям ({datetime.now().strftime('%Y-%m-%d %H:%M')})", "*"*60 + "\n"]
    grand_total = sum(sum(cat.values()) if 'СГТ' in cat else cat.get('Общее', 0) for floor_data in category_data.values() for cat in floor_data.values())
    wh_thresholds = FILL_THRESHOLDS_PER_FLOOR.get('Общая_Площадка', {}).get('Общее')
    wh_status, wh_percent, _ = get_status_by_thresholds(grand_total, wh_thresholds)
    report_lines.extend(["--- ОБЩАЯ ЗАПОЛНЕННОСТЬ СКЛАДА ---", f"  > Всего товаров: {grand_total:,.0f} шт.".replace(',', ' '), f"  > СТАТУС: {wh_status} ({wh_percent:.1f}%)\n", "--- ДЕТАЛИЗАЦИЯ ПО ЭТАЖАМ И КАТЕГОРИЯМ ---"])
    for floor, floor_data in sorted(category_data.items()):
        total_floor = sum(sum(cat.values()) if 'СГТ' in cat else cat.get('Общее', 0) for cat in floor_data.values())
        floor_status, floor_percent, _ = get_status_by_thresholds(total_floor, FILL_THRESHOLDS_PER_FLOOR.get(floor, {}).get('Общее'))
        report_lines.append(f"\n## ЭТАЖ {floor} (Всего: {total_floor:,.0f} шт. | СТАТУС: {floor_status} ({floor_percent:.1f}%))".replace(',', ' '))
        for category, cat_data in sorted(floor_data.items()):
            if category == 'Неопределенная категория' and cat_data['Общее'] == 0: continue
            cat_total = sum(cat_data.values()) if 'СГТ' in cat_data else cat_data.get('Общее', 0)
            cat_thresholds = FILL_THRESHOLDS_PER_FLOOR.get(floor, {}).get(category)
            cat_status, cat_percent, _ = get_status_by_thresholds(cat_total, cat_thresholds)
            report_lines.append(f"  - Категория: {category} (Всего: {cat_total:,.0f} шт.) - СТАТУС: {cat_status} ({cat_percent:.1f}%)".replace(',', ' '))
            if 'СГТ' in cat_data:
                report_lines.append(f"    - СГТ: {cat_data.get('СГТ', 0):,.0f} шт.".replace(',', ' ')); report_lines.append(f"    - МГТ: {cat_data.get('МГТ', 0):,.0f} шт.".replace(',', ' '))
    final_report_text = "\n".join(report_lines)
    print("\n" + final_report_text)
    try:
        with open(os.path.join(BASE_DOWNLOAD_DIR, txt_filename), 'w', encoding='utf-8') as f: f.write(final_report_text)
        print(f"\n✓ Текстовый отчет сохранен: {txt_filename}")
    except Exception as e: print(f"✗ Не удалось сохранить текстовый отчет: {e}")

def generate_detailed_excel_report(report_name, report_data):
    if not report_data: print("Нет данных для создания Excel отчета."); return
    excel_filename = f'detailed_report_{report_name}_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    df = pd.DataFrame(report_data)
    df.sort_values(by=['Этаж', 'Ряд'], inplace=True)
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for floor_num, floor_df in df.groupby('Этаж'):
        ws = wb.create_sheet(title=f"Этаж {floor_num}")
        headers = ['Ряд', 'Макс. вместимость (100%)', 'Фактическое кол-во ШК', '% заполнения', 'Осталось ШК до 95%', 'Статус']
        ws.append(headers)
        for _, row in floor_df.iterrows():
            max_fill_display = f"{row['Макс. вместимость (100%)']:,}".replace(',', ' ') if pd.notna(row['Макс. вместимость (100%)']) else 'N/A'
            actual_items_display = f"{row['Фактическое кол-во ШК']:,}".replace(',', ' ')
            percent_display = f"{row['% заполнения']:.1f}%" if pd.notna(row['% заполнения']) else 'N/A'
            items_to_95_display = f"{row['Осталось ШК до 95%']:,}".replace(',', ' ') if isinstance(row['Осталось ШК до 95%'], int) else 'N/A'
            display_row = [row['Ряд'], max_fill_display, actual_items_display, percent_display, items_to_95_display, row['Статус']]
            ws.append(display_row)
            if 'Цвет' in row and row['Цвет']:
                for cell in ws[ws.max_row]: cell.fill = row['Цвет']
        for i, _ in enumerate(headers):
            max_length = max((len(str(cell.value)) for cell in ws[get_column_letter(i + 1)] if cell.value is not None), default=12)
            ws.column_dimensions[get_column_letter(i + 1)].width = max_length + 2
    try:
        wb.save(os.path.join(BASE_DOWNLOAD_DIR, excel_filename))
        print(f"✓ Excel-отчет сохранен: {excel_filename}\n")
    except Exception as e: print(f"✗ Не удалось сохранить Excel-отчет: {e}")

def df_to_gsheet(df, worksheet_name):
    try:
        scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file(GSHEET_CRED_FILE, scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(GSHEET_ID)
        
        try:
            existing_worksheet = sh.worksheet(worksheet_name)
            sh.del_worksheet(existing_worksheet)
            print(f"Удален существующий лист '{worksheet_name}'.")
        except gspread.exceptions.WorksheetNotFound:
            pass 
        
        df_clean = df.copy()
        df_clean = df_clean.astype(str)
        worksheet = sh.add_worksheet(title=worksheet_name, rows=str(len(df_clean) + 1), cols=str(len(df_clean.columns)))
        worksheet.update([df_clean.columns.values.tolist()] + df_clean.values.tolist(), value_input_option='USER_ENTERED')
        print(f"✓ Данные записаны на вкладку '{worksheet_name}' в Google Sheets")
    except FileNotFoundError:
        print(f"ОШИБКА: Файл '{GSHEET_CRED_FILE}' не найден. Убедитесь, что он лежит рядом со скриптом.")
    except Exception as e:
        if "404" in str(e):
            print(f"ОШИБКА 404 (Не найдено) при работе с листом '{worksheet_name}'. Проверьте GSHEET_ID и права доступа.")
        else:
            print(f"ОШИБКА выгрузки в Google Sheets ({worksheet_name}): {e}")

def generate_google_sheets_reports(report_data):
    if not report_data:
        print("Нет данных для выгрузки в Google Sheets.")
        return

    df = pd.DataFrame(report_data)
    df.sort_values(by=['Этаж', 'Ряд'], inplace=True)
    
    if 'Цвет' in df.columns:
        df = df.drop(columns=['Цвет'])
        
    for floor_num, floor_df in df.groupby('Этаж'):
        display_df = floor_df.copy()
        display_df['% заполнения'] = display_df['% заполнения'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "N/A")
        cols_order = ['Ряд', 'Макс. вместимость (100%)', 'Фактическое кол-во ШК', '% заполнения', 'Осталось ШК до 95%', 'Статус']
        display_df = display_df[cols_order]
        df_to_gsheet(display_df, f"Этаж {floor_num}")

def main():
    try:
        with requests.Session() as session:
            print("Начало процесса скачивания данных...")
            token = login_and_get_token(session)
            for warehouse in WAREHOUSES_TO_PROCESS:
                download_and_save_data(session, token, warehouse, os.path.join(BASE_DOWNLOAD_DIR, warehouse['name']))
            print("Процесс скачивания данных завершен.")
    except Exception as e:
        print(f"\nОшибка при скачивании данных: {e}")
        return

    all_category_data = {}
    all_row_data_for_excel = []
    
    print("\n" + "="*60 + "\nНАЧАЛО ОБЪЕДИНЕННОГО АНАЛИЗА\n" + "="*60)
    for warehouse in WAREHOUSES_TO_PROCESS:
        wh_dir = os.path.join(BASE_DOWNLOAD_DIR, warehouse['name'])
        category_data, row_data = analyze_warehouse_data(wh_dir)
        for floor, data in category_data.items():
            if floor not in all_category_data:
                all_category_data[floor] = data.copy()
            else:
                for key, val in data.items():
                    if key in all_category_data[floor] and isinstance(val, dict):
                        for sub_key, sub_val in val.items():
                            all_category_data[floor][key][sub_key] = all_category_data[floor][key].get(sub_key, 0) + sub_val
                    else:
                        all_category_data[floor][key] = val.copy()
        all_row_data_for_excel.extend(row_data)

    if all_category_data or all_row_data_for_excel:
        print("\n" + "="*60 + "\nГЕНЕРАЦИЯ СВОДНЫХ ОТЧЕТОВ\n" + "="*60)
        final_report_name = "Сводный_Отчет_по_Всем_Складам"
        generate_summary_txt_report(final_report_name, all_category_data)
        generate_detailed_excel_report(final_report_name, all_row_data_for_excel)
        print("\n" + "-"*20 + " ВЫГРУЗКА В GOOGLE SHEETS " + "-"*20)
        generate_google_sheets_reports(all_row_data_for_excel)
    else:
        print("Не найдено данных для создания отчетов.")

    print("\n" + "="*60 + "\nОЧИСТКА ИСХОДНЫХ ФАЙЛОВ\n" + "="*60)
    for warehouse in WAREHOUSES_TO_PROCESS:
        wh_dir_to_delete = os.path.join(BASE_DOWNLOAD_DIR, warehouse['name'])
        try:
            if os.path.isdir(wh_dir_to_delete):
                shutil.rmtree(wh_dir_to_delete)
                print(f"✓ Папка '{wh_dir_to_delete}' и ее содержимое удалены.")
        except OSError as e:
            print(f"✗ Не удалось удалить папку '{wh_dir_to_delete}': {e}")
            
    print("\nРабота скрипта полностью завершена.")

if __name__ == '__main__':
    main()
